"""Unit 118: seminar end-to-end acceptance / regression.

Closes out the seminar feature (Units 115-117) by driving one real unit through
the whole admin path — create -> allocate -> schedule -> capacity -> solve ->
export — through the actual services, solver, and export renderer (no new
behaviour of its own). It proves the three slices compose:

* hidden allocations give each enrolled student exactly one tutorial (a
  balanced partition) while every seminar session, like a lecture, holds the
  whole enrolled cohort (Unit 115);
* the export renders independent ``Tutorial A/B`` and ``Seminar A/B`` letter
  series that agree with the frontend on-card letters (Units 116/117);
* capacity is checked against the seminar's group, which is now the full
  enrolment; and
* the solver schedules unscheduled seminars exactly like tutorials, with no
  feasibility regression and deterministic output.

Uses the in-memory SQLite ``db`` fixture from conftest.
"""
import itertools
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pytest

from api.errors import AppError
from models.lecturer import AvailabilityDay, AvailabilitySlot, Lecturer, LecturerTitle
from models.room import Room, RoomType
from models.session import Session, SessionType
from models.session_allocation import SessionStudentAllocation
from models.student import Student
from schemas.assignment import AssignmentItem, AssignmentSaveRequest
from schemas.session import SessionCreate
from schemas.unit import UnitCreate
from services.assignment import save_assignments
from services.session import create_session
from services.unit import create_unit
from services import timetable_excel_export as ex
from solver.model import solve_timetable
from solver.snapshot import build_solver_input_snapshot

import io

import openpyxl


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------

_student_numbers = itertools.count(40_000_000)


def make_lecturer(db, lid="lec1", first="Luke", last="Holohan") -> Lecturer:
    lec = Lecturer(id=lid, title=LecturerTitle.DR, first_name=first, last_name=last)
    db.add(lec)
    db.commit()
    return lec


def make_students(db, count, year_level=1) -> list[str]:
    ids = []
    for i in range(count):
        sid = f"s{i:02d}"
        db.add(
            Student(
                id=sid,
                student_number=str(next(_student_numbers)),
                first_name="Stu",
                last_name=sid,
                year_level=year_level,
            )
        )
        ids.append(sid)
    db.commit()
    return ids


def base_rooms(db, capacity=100) -> dict[str, Room]:
    """Create the eight fixed-order template rooms by name (export-compatible)."""
    rooms = {}
    for i, name in enumerate(ex.ROOM_ORDER):
        r = Room(id=f"r-{i}", name=name, capacity=capacity, room_type=RoomType.LECTURE)
        db.add(r)
        rooms[name] = r
    db.commit()
    return rooms


def sessions_of(db, unit_id, session_type) -> list[Session]:
    return list(
        db.query(Session).filter(
            Session.unit_id == unit_id, Session.session_type == session_type
        )
    )


def group_map(db, unit_id, session_type) -> dict[str, str]:
    """student_id -> session_id for the unit's sessions of a given type."""
    ids = {s.id for s in sessions_of(db, unit_id, session_type)}
    mapping: dict[str, str] = {}
    for row in db.query(SessionStudentAllocation).filter(
        SessionStudentAllocation.session_id.in_(ids or {"__none__"})
    ):
        mapping[row.student_id] = row.session_id
    return mapping


def group_sizes(db, unit_id, session_type) -> list[int]:
    sizes = []
    for s in sessions_of(db, unit_id, session_type):
        sizes.append(
            db.query(SessionStudentAllocation)
            .filter(SessionStudentAllocation.session_id == s.id)
            .count()
        )
    return sorted(sizes)


def make_unit(db, n_students, n_tutorials, n_seminars, code="HIS101"):
    make_lecturer(db)
    students = make_students(db, n_students)
    unit = create_unit(
        db,
        UnitCreate(code=code, name="History", lecturer_ids=["lec1"], student_ids=students),
    )
    for _ in range(n_tutorials):
        create_session(db, unit.id, SessionCreate(session_type=SessionType.TUTORIAL, duration=1))
    for _ in range(n_seminars):
        create_session(db, unit.id, SessionCreate(session_type=SessionType.SEMINAR, duration=1))
    return unit, students


def load_ws(stream: io.BytesIO):
    wb = openpyxl.load_workbook(stream)
    return wb.worksheets[0]


# ---------------------------------------------------------------------------
# Step 2-3: allocation — one-of-each, balanced, independent partitions
# ---------------------------------------------------------------------------


def test_each_student_has_exactly_one_tutorial_and_every_seminar(db):
    unit, students = make_unit(db, 10, n_tutorials=3, n_seminars=2)
    tut_map = group_map(db, unit.id, SessionType.TUTORIAL)
    assert set(tut_map.keys()) == set(students)
    # Every seminar session holds the whole cohort, like a lecture.
    for sem in sessions_of(db, unit.id, SessionType.SEMINAR):
        members = {
            r.student_id
            for r in db.query(SessionStudentAllocation).filter(
                SessionStudentAllocation.session_id == sem.id
            )
        }
        assert members == set(students)


def test_tutorials_balanced_and_seminars_hold_full_cohort(db):
    unit, _ = make_unit(db, 10, n_tutorials=3, n_seminars=4)
    # Tutorials: 10 across 3 -> 4/3/3 (balanced partition). Seminars: every
    # session holds the full cohort of 10, like lectures.
    assert group_sizes(db, unit.id, SessionType.TUTORIAL) == [3, 3, 4]
    assert group_sizes(db, unit.id, SessionType.SEMINAR) == [10, 10, 10, 10]


def test_partitions_are_independent_both_directions(db):
    unit, _ = make_unit(db, 8, n_tutorials=2, n_seminars=2)
    tut_before = group_map(db, unit.id, SessionType.TUTORIAL)

    # Adding a seminar rebalances seminars only; the tutorial partition is untouched.
    create_session(db, unit.id, SessionCreate(session_type=SessionType.SEMINAR, duration=1))
    assert group_map(db, unit.id, SessionType.TUTORIAL) == tut_before

    # Adding a tutorial rebalances tutorials only; the (now 3-group) seminar
    # partition is untouched.
    sem_baseline = group_map(db, unit.id, SessionType.SEMINAR)
    create_session(db, unit.id, SessionCreate(session_type=SessionType.TUTORIAL, duration=1))
    assert group_map(db, unit.id, SessionType.SEMINAR) == sem_baseline


# ---------------------------------------------------------------------------
# Step 5: capacity fires against the seminar's group, now the full enrolment
# ---------------------------------------------------------------------------


def _save_one(db, session_id, room_id, slot=AvailabilitySlot.S1, day=AvailabilityDay.MONDAY):
    return save_assignments(
        db,
        AssignmentSaveRequest(
            assignments=[
                AssignmentItem(session_id=session_id, day=day, start_slot=slot, room_id=room_id)
            ]
        ),
    )


def test_seminar_capacity_uses_full_enrolment(db):
    unit, _ = make_unit(db, 10, n_tutorials=0, n_seminars=2)  # two seminars of 10
    base_rooms(db, capacity=100)
    # A seminar now holds the whole cohort, so capacity is checked against 10.
    db.add(Room(id="exact", name="exact", capacity=10, room_type=RoomType.TUTORIAL))
    db.add(Room(id="tight", name="tight", capacity=9, room_type=RoomType.TUTORIAL))
    db.commit()
    sem = sessions_of(db, unit.id, SessionType.SEMINAR)[0]

    # capacity == cohort size (10) succeeds.
    saved = _save_one(db, sem.id, "exact")
    assert saved[0].student_count == 10

    # capacity < cohort size (9 < 10) is rejected by the defensive capacity check.
    with pytest.raises(AppError):
        _save_one(db, sem.id, "tight")


# ---------------------------------------------------------------------------
# Step 6: solver schedules unscheduled seminars like tutorials, deterministically
# ---------------------------------------------------------------------------


def test_solver_schedules_seminars_like_tutorials_and_is_deterministic(db):
    unit, _ = make_unit(db, 6, n_tutorials=1, n_seminars=1)
    base_rooms(db, capacity=100)  # eight rooms, plenty of feasible cells

    snapshot = build_solver_input_snapshot(db)
    # Both the tutorial and the seminar are unscheduled variables for the solver.
    assert len(snapshot.unscheduled_session_ids) == 2

    result = solve_timetable(snapshot)
    assert result.scheduled_count == 2
    assert result.unscheduled_count == 0

    # Deterministic: a second solve of the same snapshot yields identical placements.
    again = solve_timetable(build_solver_input_snapshot(db))
    placements = lambda r: sorted(
        (g.session_id, g.day, g.start_slot, g.room_id) for g in r.generated_assignments
    )
    assert placements(result) == placements(again)


# ---------------------------------------------------------------------------
# Step 4 + 7: export renders independent Tutorial/Seminar letters that agree
# with the on-card letters, end to end through the real allocation + save path
# ---------------------------------------------------------------------------


def test_export_shows_independent_tutorial_and_seminar_letters(db):
    unit, _ = make_unit(db, 8, n_tutorials=2, n_seminars=2, code="THE202")
    rooms = base_rooms(db, capacity=100)

    tuts = sorted(sessions_of(db, unit.id, SessionType.TUTORIAL), key=lambda s: s.id)
    sems = sorted(sessions_of(db, unit.id, SessionType.SEMINAR), key=lambda s: s.id)

    # Tutorials in L1.05 (col C block); seminars in PDS (col B block); Mon & Tue.
    _save_one(db, tuts[0].id, rooms["L1.05"].id, day=AvailabilityDay.MONDAY)
    save_assignments(
        db,
        AssignmentSaveRequest(
            assignments=[
                AssignmentItem(session_id=tuts[0].id, day=AvailabilityDay.MONDAY, start_slot=AvailabilitySlot.S1, room_id=rooms["L1.05"].id),
                AssignmentItem(session_id=tuts[1].id, day=AvailabilityDay.TUESDAY, start_slot=AvailabilitySlot.S1, room_id=rooms["L1.05"].id),
                AssignmentItem(session_id=sems[0].id, day=AvailabilityDay.MONDAY, start_slot=AvailabilitySlot.S1, room_id=rooms["PDS"].id),
                AssignmentItem(session_id=sems[1].id, day=AvailabilityDay.TUESDAY, start_slot=AvailabilitySlot.S1, room_id=rooms["PDS"].id),
            ]
        ),
    )

    ws = load_ws(ex.generate_timetable_export(db, "Acceptance"))
    # Independent A/B series: tutorials and seminars each start at A in timetable
    # order and never share a counter (row 6 == slot s1).
    assert ws["C6"].value == "THE202 Tutorial A (LH)"   # Monday L1.05
    assert ws["K6"].value == "THE202 Tutorial B (LH)"   # Tuesday L1.05
    assert ws["B6"].value == "THE202 Seminar A (LH)"    # Monday PDS
    assert ws["J6"].value == "THE202 Seminar B (LH)"    # Tuesday PDS
