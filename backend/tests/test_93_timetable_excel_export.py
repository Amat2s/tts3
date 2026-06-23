"""Tests for Unit 93: backend timetable Excel export template and API.

Covers the export service (template load, fixed day/room/slot mapping, one- and
multi-slot class rendering, subject styling, session labels, tutorial lettering,
lecturer/tutor key generation, block rendering + rectangle merging, unknown-room
failure, and the no-persistence guarantee) plus the protected route (auth, title
validation, and the ``.xlsx`` streaming response). Uses the in-memory SQLite
``db`` fixture from conftest.
"""
import asyncio
import io
import json
import os
import sys
from datetime import date
from urllib.parse import urlencode

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from auth.deps import CurrentAdmin, get_current_admin
from db.deps import get_db
from main import app
from models.assignment import TimetableAssignment
from models.lecturer import AvailabilityDay, AvailabilitySlot, Lecturer, LecturerTitle
from models.room import Room, RoomType
from models.session import Session, SessionType
from models.timetable_block import BlockColour, TimetableBlockCell, TimetableBlockGroup
from models.unit import Unit
from api.errors import AppError
from services import timetable_excel_export as ex


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------


def make_room(db, room_id, name, capacity=100, room_type=RoomType.LECTURE) -> Room:
    r = Room(id=room_id, name=name, capacity=capacity, room_type=room_type)
    db.add(r)
    db.commit()
    return r


def make_unit(db, unit_id="u1", code="HIS101", year_level=1) -> Unit:
    u = Unit(id=unit_id, code=code, name=code, year_level=year_level)
    db.add(u)
    db.commit()
    return u


def make_lecturer(db, lid="l1", first="Steve", last="Chavura", title=LecturerTitle.DR) -> Lecturer:
    lec = Lecturer(id=lid, title=title, first_name=first, last_name=last)
    db.add(lec)
    db.commit()
    return lec


def make_session(db, sid, unit, lecturer, session_type=SessionType.LECTURE, duration=1) -> Session:
    s = Session(
        id=sid,
        unit_id=unit.id,
        session_type=session_type,
        duration=duration,
        lecturer_id=lecturer.id if lecturer else None,
    )
    db.add(s)
    db.commit()
    return s


def make_assignment(db, aid, session, day, slot, room) -> TimetableAssignment:
    a = TimetableAssignment(
        id=aid, session_id=session.id, day=day, start_slot=slot, room_id=room.id
    )
    db.add(a)
    db.commit()
    return a


def make_block(db, gid, name, colour, cells) -> TimetableBlockGroup:
    g = TimetableBlockGroup(id=gid, name=name, colour=colour)
    db.add(g)
    db.commit()
    for i, (day, slot, room) in enumerate(cells):
        db.add(
            TimetableBlockCell(
                id=f"{gid}-c{i}",
                block_group_id=gid,
                day=day,
                slot=slot,
                room_id=room.id,
            )
        )
    db.commit()
    return g


def base_rooms(db) -> dict[str, Room]:
    """Create the eight fixed-order template rooms by name."""
    return {
        name: make_room(db, f"r-{i}", name)
        for i, name in enumerate(ex.ROOM_ORDER)
    }


def load_ws(stream: io.BytesIO):
    wb = openpyxl.load_workbook(stream)
    return wb, wb.worksheets[0]


def coord(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def merges(ws) -> set[str]:
    return {str(m) for m in ws.merged_cells.ranges}


# ---------------------------------------------------------------------------
# ASGI test client (supports query strings + response headers)
# ---------------------------------------------------------------------------


class ASGIResponse:
    def __init__(self, status_code, body, headers):
        self.status_code = status_code
        self.body = body
        self.headers = headers

    def json(self):
        return json.loads(self.body.decode("utf-8"))


class ASGITestClient:
    def __init__(self, application):
        self.application = application

    def get(self, path, params=None):
        return asyncio.run(self._request("GET", path, params))

    async def _request(self, method, path, params):
        query_string = urlencode(params or {}).encode("ascii")
        scope = {
            "type": "http",
            "asgi": {"version": "3.0"},
            "http_version": "1.1",
            "method": method,
            "scheme": "http",
            "path": path,
            "raw_path": path.encode("ascii"),
            "query_string": query_string,
            "headers": [(b"host", b"testserver")],
            "client": ("testclient", 50000),
            "server": ("testserver", 80),
        }
        messages = []
        received = False
        never = asyncio.Event()

        async def receive():
            nonlocal received
            if not received:
                received = True
                return {"type": "http.request", "body": b"", "more_body": False}
            # After the request is delivered, block (yielding to the event loop)
            # so StreamingResponse's disconnect listener does not cancel the body
            # stream early. The framework cancels this await once the response is
            # fully sent, which ends the request cleanly.
            await never.wait()
            return {"type": "http.disconnect"}

        async def send(message):
            messages.append(message)

        await self.application(scope, receive, send)
        start = next(m for m in messages if m["type"] == "http.response.start")
        body = b"".join(
            m.get("body", b"") for m in messages if m["type"] == "http.response.body"
        )
        headers = {
            k.decode("latin-1").lower(): v.decode("latin-1")
            for k, v in start["headers"]
        }
        return ASGIResponse(start["status"], body, headers)


@pytest.fixture
def client(db):
    def override_db():
        yield db

    def override_admin():
        return CurrentAdmin(user_id="admin-1", email="admin@example.com")

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_admin] = override_admin
    yield ASGITestClient(app)
    app.dependency_overrides.clear()


@pytest.fixture
def unauthenticated_client(db):
    def override_db():
        yield db

    app.dependency_overrides[get_db] = override_db
    yield ASGITestClient(app)
    app.dependency_overrides.clear()


# ---------------------------------------------------------------------------
# Route: auth + title + response shape
# ---------------------------------------------------------------------------


def test_export_requires_auth(unauthenticated_client):
    res = unauthenticated_client.get("/timetable/export.xlsx", {"title": "T"})
    assert res.status_code == 401


def test_title_required_when_missing(client):
    res = client.get("/timetable/export.xlsx")
    assert res.status_code == 422
    assert res.json()["error"]["code"] == "invalid_export_title"


def test_title_required_when_blank(client):
    res = client.get("/timetable/export.xlsx", {"title": "   "})
    assert res.status_code == 422
    assert res.json()["error"]["code"] == "invalid_export_title"


def test_response_is_xlsx(client):
    res = client.get("/timetable/export.xlsx", {"title": "Semester 2, 2026 Timetable"})
    assert res.status_code == 200
    assert res.headers["content-type"].startswith(ex.XLSX_MEDIA_TYPE)
    disp = res.headers["content-disposition"]
    assert "attachment" in disp
    assert "semester-2-2026-timetable-" in disp
    assert disp.endswith('.xlsx"')
    # Body is a real, loadable workbook with a single sheet.
    wb = openpyxl.load_workbook(io.BytesIO(res.body))
    assert len(wb.worksheets) == 1


# ---------------------------------------------------------------------------
# Service: template + mapping
# ---------------------------------------------------------------------------


def test_template_loads_valid_single_sheet_workbook(db):
    stream = ex.generate_timetable_export(db, "Empty Timetable")
    wb, ws = load_ws(stream)
    assert len(wb.worksheets) == 1
    # Static content preserved; title + version set.
    assert ws["B2"].value == "Empty Timetable"
    assert ws["K29"].value == "Version 1"
    assert ws["B12"].value == "Mass/Lunch"
    assert ws["Z22"].value == "FORMAL HALL"


def test_one_slot_session_fills_both_rows(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db)
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    # s1 PDS Monday -> B6:B7 (both visual rows of the band).
    assert "B6:B7" in merges(ws)
    assert ws["B6"].value == "HIS101 Lecture (SC)"


def test_multi_slot_session_spans_all_rows(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db)
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=3)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    # 3 slots from s1 -> rows 6..11 (s1+s2+s3 bands).
    assert "B6:B11" in merges(ws)
    assert ws["B6"].value == "HIS101 Lecture (SC)"


@pytest.mark.parametrize(
    "day,room_name,slot,expected",
    [
        (AvailabilityDay.MONDAY, "PDS", AvailabilitySlot.S1, "B6"),
        (AvailabilityDay.TUESDAY, "L1.05", AvailabilitySlot.S3, "K10"),
        (AvailabilityDay.WEDNESDAY, "JTW", AvailabilitySlot.S4, "Y14"),
        (AvailabilityDay.FRIDAY, "PDS", AvailabilitySlot.S7, "AH20"),
        (AvailabilityDay.THURSDAY, "Dawson", AvailabilitySlot.S5, "AD16"),
    ],
)
def test_fixed_room_mapping(db, day, room_name, slot, expected):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db)
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, day, slot, rooms[room_name])

    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    assert ws[expected].value == "HIS101 Lecture (SC)"


def test_cell_anchor_helper():
    assert ex.cell_anchor("Monday", "PDS", "s1") == (6, 2)
    assert ex.cell_anchor("Friday", "JTW", "s7") == (20, 41)


def test_unknown_room_fails_export(db):
    # A room whose name is not in the fixed template room order.
    bad_room = make_room(db, "r-bad", "Grand Hall")
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db)
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, bad_room)

    with pytest.raises(AppError) as exc:
        ex.generate_timetable_export(db, "T")
    assert exc.value.code == "export_room_not_in_template"
    assert exc.value.status_code == 422


# ---------------------------------------------------------------------------
# Labels, tutorial lettering, lecturer key
# ---------------------------------------------------------------------------


def test_lecture_label_format(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db, first="Steve", last="Chavura")
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    assert ws["B6"].value == "HIS101 Lecture (SC)"


def test_tutorial_letters_by_timetable_order(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="THE202")
    lec = make_lecturer(db, first="Luke", last="Holohan")
    # Two tutorials of the same unit: one Tuesday s1, one Monday s1.
    t_tue = make_session(db, "t-tue", unit, lec, SessionType.TUTORIAL, duration=1)
    t_mon = make_session(db, "t-mon", unit, lec, SessionType.TUTORIAL, duration=1)
    make_assignment(db, "a-tue", t_tue, AvailabilityDay.TUESDAY, AvailabilitySlot.S1, rooms["PDS"])
    make_assignment(db, "a-mon", t_mon, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    # Monday sorts before Tuesday -> letter A on Monday, B on Tuesday.
    assert ws["B6"].value == "THE202 Tutorial A (LH)"   # Monday PDS s1
    assert ws["J6"].value == "THE202 Tutorial B (LH)"   # Tuesday PDS s1


def test_unscheduled_tutorials_do_not_consume_letters(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="THE202")
    lec = make_lecturer(db, first="Luke", last="Holohan")
    # One scheduled tutorial + one unscheduled (no assignment) tutorial.
    scheduled = make_session(db, "t-sched", unit, lec, SessionType.TUTORIAL, duration=1)
    make_session(db, "t-unsched", unit, lec, SessionType.TUTORIAL, duration=1)
    make_assignment(db, "a1", scheduled, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    # The single exported tutorial is "A"; the unscheduled one consumes nothing.
    assert ws["B6"].value == "THE202 Tutorial A (LH)"


def test_lecturer_initials_helper():
    assert ex.lecturer_initials("Steve", "Chavura") == "SC"
    assert ex.lecturer_initials("luke", "holohan") == "LH"


def test_lecturer_key_generation_and_sort(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    sc = make_lecturer(db, lid="l-sc", first="Steve", last="Chavura")
    lh = make_lecturer(db, lid="l-lh", first="Luke", last="Holohan")
    s1 = make_session(db, "s1", unit, sc, SessionType.LECTURE, duration=1)
    s2 = make_session(db, "s2", unit, lh, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s1, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])
    make_assignment(db, "a2", s2, AvailabilityDay.TUESDAY, AvailabilitySlot.S1, rooms["PDS"])

    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    # Sorted by initials (LH < SC); balanced across the two key columns.
    assert ws["A29"].value == "LH: Dr Luke Holohan"
    assert ws["D29"].value == "SC: Dr Steve Chavura"


def test_session_missing_lecturer_fails_export(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    s = make_session(db, "s1", unit, None, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    with pytest.raises(AppError) as exc:
        ex.generate_timetable_export(db, "T")
    assert exc.value.code == "export_session_missing_lecturer"


# ---------------------------------------------------------------------------
# Blocks
# ---------------------------------------------------------------------------


def test_named_block_exports_name_with_block_style(db):
    rooms = base_rooms(db)
    make_block(
        db, "g1", "Chapel", BlockColour.GOLD,
        [(AvailabilityDay.MONDAY, AvailabilitySlot.S5, rooms["PDS"])],
    )
    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    # s5 Monday PDS -> rows 16-17, col B.
    assert "B16:B17" in merges(ws)
    assert ws["B16"].value == "Chapel"
    # Template-derived gold event fill, not the grey empty-cell fill.
    assert ws["B16"].fill.fgColor.rgb == "FFcc9900"


def test_unnamed_block_exports_blank_with_block_style(db):
    rooms = base_rooms(db)
    make_block(
        db, "g1", None, None,
        [(AvailabilityDay.MONDAY, AvailabilitySlot.S5, rooms["PDS"])],
    )
    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    assert "B16:B17" in merges(ws)
    assert ws["B16"].value is None
    # Unnamed blocks use the template's grey "blocked" (empty-cell) fill.
    assert ws["B16"].fill.fgColor.rgb == "FFd9d9d9"


def test_block_rectangle_merging(db):
    rooms = base_rooms(db)
    # A 2-room x 2-slot rectangle: PDS+L1.05 across s4+s5 on Monday.
    cells = [
        (AvailabilityDay.MONDAY, AvailabilitySlot.S4, rooms["PDS"]),
        (AvailabilityDay.MONDAY, AvailabilitySlot.S4, rooms["L1.05"]),
        (AvailabilityDay.MONDAY, AvailabilitySlot.S5, rooms["PDS"]),
        (AvailabilityDay.MONDAY, AvailabilitySlot.S5, rooms["L1.05"]),
    ]
    make_block(db, "g1", "Staff Meeting", BlockColour.LIGHT_BLUE, cells)

    _, ws = load_ws(ex.generate_timetable_export(db, "T"))
    # s4 top row 14, s5 bottom row 17; cols B..C -> single merged rectangle.
    assert "B14:C17" in merges(ws)
    assert ws["B14"].value == "Staff Meeting"


def test_block_room_not_in_template_fails(db):
    bad_room = make_room(db, "r-bad", "Grand Hall")
    make_block(
        db, "g1", "Event", BlockColour.GOLD,
        [(AvailabilityDay.MONDAY, AvailabilitySlot.S5, bad_room)],
    )
    with pytest.raises(AppError) as exc:
        ex.generate_timetable_export(db, "T")
    assert exc.value.code == "export_room_not_in_template"


# ---------------------------------------------------------------------------
# No persistence
# ---------------------------------------------------------------------------


def test_export_persists_nothing(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db)
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    before_assignments = db.query(TimetableAssignment).count()
    before_blocks = db.query(TimetableBlockGroup).count()

    stream = ex.generate_timetable_export(db, "T")

    # Returns an in-memory stream and leaves saved state untouched.
    assert isinstance(stream, io.BytesIO)
    assert db.query(TimetableAssignment).count() == before_assignments
    assert db.query(TimetableBlockGroup).count() == before_blocks


def test_export_filename_slug():
    assert (
        ex.export_filename("Semester 2, 2026 Timetable", date(2026, 6, 20))
        == "semester-2-2026-timetable-2026-06-20.xlsx"
    )
    # Falls back to a stable slug when the title has no alphanumerics.
    assert ex.export_filename("!!!", date(2026, 1, 1)) == "timetable-2026-01-01.xlsx"
