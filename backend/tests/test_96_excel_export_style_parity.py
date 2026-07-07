"""Unit 96/97: Excel export styling parity tests.

These prove *style* parity (not just placement), comparing the exported workbook
against the canonical source sheet ``S2, 2025 Timetable `` (trailing space) of
the real pristine template ``assets/excel/Timetable S2 2025 DRAFT.xlsx`` — the
visual contract. They check that the export preserves the template's columns,
row heights, sheet-format properties, page setup, tab colour, sheet view zoom,
margins, static merged ranges, and static cell styles/values verbatim; that the
blanked grid keeps the per-day empty fills (Tuesday/Thursday white, Mon/Wed/Fri
grey) with bold dark borders; and that dynamic class and block cells are painted
with theme-based styles copied from the template's own class/event exemplar
cells (never frontend subject tokens or approximated RGB).
"""
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string, get_column_letter

from models.lecturer import AvailabilityDay, AvailabilitySlot, LecturerTitle
from models.session import SessionType
from models.timetable_block import BlockColour
from services import timetable_excel_export as ex

# Reuse the fixture builders from the Unit 93 suite.
from tests.test_93_timetable_excel_export import (
    base_rooms,
    load_ws,
    make_assignment,
    make_block,
    make_lecturer,
    make_session,
    make_unit,
    merges,
)

# Canonical source template — the visual contract (Unit 97: the real pristine
# template, no trailing space; the canonical sheet is "S2, 2025 Timetable "
# WITH a trailing space, not worksheets[0] which is a generic template sheet).
TEMPLATE_SOURCE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "assets", "excel", "Timetable S2 2025 DRAFT.xlsx",
)
TEMPLATE_SOURCE_SHEET = "S2, 2025 Timetable "

# Grid (dynamic) region: every row inside an s1-s7 slot band, columns B..AO.
SLOT_TOP_ROWS = [6, 8, 10, 14, 16, 18, 20]
GRID_ROWS = {r for top in SLOT_TOP_ROWS for r in (top, top + 1)}
GRID_MIN_COL, GRID_MAX_COL = 2, 41


@pytest.fixture(scope="module")
def template_ws():
    wb = openpyxl.load_workbook(TEMPLATE_SOURCE_PATH)
    return wb[TEMPLATE_SOURCE_SHEET]


def export(db, title="Parity Timetable"):
    return load_ws(ex.generate_timetable_export(db, title))[1]


# --- Style signature helpers -------------------------------------------------


def _color_sig(c):
    if c is None:
        return None
    return (
        c.rgb if isinstance(c.rgb, str) else None,
        c.theme if isinstance(c.theme, int) else None,
        round(c.tint, 6) if c.tint else 0.0,
        c.indexed if isinstance(c.indexed, int) else None,
    )


def _side_sig(s):
    return (s.style, _color_sig(s.color))


def style_sig(cell):
    """A fully comparable signature of a cell's visual style."""
    f, fill, b, a = cell.font, cell.fill, cell.border, cell.alignment
    return {
        "font": (f.name, f.size, bool(f.b), bool(f.i), f.u, _color_sig(f.color)),
        "fill": (fill.patternType, _color_sig(fill.fgColor) if fill.patternType else None),
        "border": tuple(_side_sig(s) for s in (b.top, b.bottom, b.left, b.right)),
        "align": (a.horizontal, a.vertical, bool(a.wrap_text)),
        "numfmt": cell.number_format,
    }


def is_grid_merge(rng) -> bool:
    rows = set(range(rng.min_row, rng.max_row + 1))
    return (
        rows <= GRID_ROWS
        and rng.min_col >= GRID_MIN_COL
        and rng.max_col <= GRID_MAX_COL
    )


def static_merges(ws) -> set[str]:
    return {str(m) for m in ws.merged_cells.ranges if not is_grid_merge(m)}


# ---------------------------------------------------------------------------
# Structural parity
# ---------------------------------------------------------------------------


def test_one_sheet_with_template_name(db, template_ws):
    wb, ws = load_ws(ex.generate_timetable_export(db, "T"))
    assert len(wb.worksheets) == 1
    assert ws.title == template_ws.title


def test_dimension_is_a1_ap40(db):
    assert export(db).dimensions == "A1:AP40"


def test_column_dims_match_template(db, template_ws):
    ws = export(db)
    for idx in range(1, 43):  # A..AP
        letter = get_column_letter(idx)
        t = template_ws.column_dimensions[letter]
        g = ws.column_dimensions[letter]
        assert (round(g.width, 6) if g.width else None) == (
            round(t.width, 6) if t.width else None
        ), f"col {letter} width"
        assert bool(g.customWidth) == bool(t.customWidth), f"col {letter} customWidth"
        assert bool(g.bestFit) == bool(t.bestFit), f"col {letter} bestFit"


def test_no_autofit_changed_column_widths(db, template_ws):
    """The export must not recompute/autofit widths: every column width is the
    template's verbatim, never a freshly-measured value."""
    ws = export(db)
    template_widths = [
        round(template_ws.column_dimensions[get_column_letter(i)].width, 6)
        for i in range(1, 43)
    ]
    export_widths = [
        round(ws.column_dimensions[get_column_letter(i)].width, 6)
        for i in range(1, 43)
    ]
    assert export_widths == template_widths


def test_row_heights_match_template(db, template_ws):
    ws = export(db)
    for r in range(1, 41):
        t = template_ws.row_dimensions[r]
        g = ws.row_dimensions[r]
        assert (round(g.height, 6) if g.height else None) == (
            round(t.height, 6) if t.height else None
        ), f"row {r} height"
        assert bool(g.customHeight) == bool(t.customHeight), f"row {r} customHeight"


def test_sheet_format_properties_match_template(db, template_ws):
    ws = export(db)
    t, g = template_ws.sheet_format, ws.sheet_format
    assert g.defaultRowHeight == t.defaultRowHeight
    assert g.baseColWidth == t.baseColWidth
    assert g.defaultColWidth == t.defaultColWidth


def test_page_margins_match_template(db, template_ws):
    ws = export(db)
    t, g = template_ws.page_margins, ws.page_margins
    for attr in ("left", "right", "top", "bottom", "header", "footer"):
        assert getattr(g, attr) == getattr(t, attr), attr


def test_page_setup_matches_template(db, template_ws):
    """Landscape, page scale 35, and fitToPage carry over from the template."""
    ws = export(db)
    assert ws.page_setup.orientation == template_ws.page_setup.orientation == "landscape"
    assert ws.page_setup.scale == template_ws.page_setup.scale == 35
    assert ws.sheet_properties.pageSetUpPr is not None
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True


def test_tab_colour_matches_template(db, template_ws):
    ws = export(db)
    assert ws.sheet_properties.tabColor.rgb == template_ws.sheet_properties.tabColor.rgb == "FF00B050"


def test_sheet_view_zoom_matches_template(db, template_ws):
    ws = export(db)
    assert ws.sheet_view.zoomScale == template_ws.sheet_view.zoomScale == 90


def test_default_row_height_matches_template(db, template_ws):
    ws = export(db)
    assert ws.sheet_format.defaultRowHeight == template_ws.sheet_format.defaultRowHeight == 12.95


def test_static_merged_ranges_match_template(db, template_ws):
    ws = export(db)
    assert static_merges(ws) == static_merges(template_ws)


# ---------------------------------------------------------------------------
# Per-day empty-cell fills + bold borders (the blanked grid)
# ---------------------------------------------------------------------------


def _fill_daytype(cell):
    """(theme, tint) of a theme-based empty grid fill."""
    fg = cell.fill.fgColor
    return (fg.theme, round(fg.tint, 4) if fg.tint else 0.0)


# Representative empty grid cells, one interior cell per day block. Tuesday
# (J-Q) and Thursday (Z-AG) are white (theme0 tint 0.0); Monday/Wednesday/Friday
# are grey (theme0 tint -0.15). These coords are empty in the source too.
GREY_EMPTY = ["D6", "E6", "W6", "AH6"]      # Mon / Mon / Wed / Fri
WHITE_EMPTY = ["K6", "L6", "AA6", "AB8"]    # Tue / Tue / Thu / Thu


@pytest.mark.parametrize("coord", GREY_EMPTY)
def test_empty_grid_cells_grey_on_mon_wed_fri(db, coord):
    ws = export(db)
    assert _fill_daytype(ws[coord]) == (0, -0.15), coord


@pytest.mark.parametrize("coord", WHITE_EMPTY)
def test_empty_grid_cells_white_on_tue_thu(db, coord):
    ws = export(db)
    assert _fill_daytype(ws[coord]) == (0, 0.0), coord


def test_every_grid_box_is_uniform_medium_box(db):
    """Every grid box (all rooms, all slots, all five days) has an identical,
    clearly-bordered look: a medium (bold) box on all four sides in the dark
    indexed-64 colour — never the old thin light-grey (c6c6c6) interior — and
    the correct day-type fill (Tue/Thu white, Mon/Wed/Fri grey). This is the
    uniform-border contract: adjacent room columns and slot rows all read the
    same, and no cell is left unfilled (e.g. the old T16-17 white gap)."""
    ws = export(db)
    white_starts = {10, 26}  # Tuesday (J-Q), Thursday (Z-AG)
    for start in (2, 10, 18, 26, 34):
        expected = (0, 0.0) if start in white_starts else (0, -0.15)
        for col in range(start, start + 8):
            for top in SLOT_TOP_ROWS:
                cell = ws.cell(row=top, column=col)
                coord = cell.coordinate
                assert _fill_daytype(cell) == expected, coord
                b = cell.border
                # Merged 2-row band: the anchor owns top/left/right; bottom is
                # medium on the band's bottom cell. Check the box is medium and
                # dark on the visible edges.
                assert b.top.style == "medium", f"{coord} top"
                assert b.left.style == "medium", f"{coord} left"
                assert b.right.style == "medium", f"{coord} right"
                for side in (b.top, b.left, b.right):
                    assert side.color is None or side.color.rgb != "FFc6c6c6", coord
                bottom = ws.cell(row=top + 1, column=col)
                assert bottom.border.bottom.style == "medium", f"{coord} bottom"


# ---------------------------------------------------------------------------
# Static cell style + value parity
# ---------------------------------------------------------------------------


# (coordinate, expected static value or None to skip the value check)
STATIC_STYLE_CELLS = [
    ("B4", "Monday"),       # day header
    ("AH4", "Friday"),      # day header (last day block)
    ("B5", "PDS"),          # room header
    ("G5", "L1.10"),        # room header
    ("A6", "9.00 AM - 9.50 AM"),   # time label
    ("A12", "12.00 PM - 1.30PM"),  # lunch time label
    ("B12", "Mass/Lunch"),  # static lunch row
    ("AH12", "Mass/Lunch"),
    ("Z22", "FORMAL HALL"), # static event banner
    ("AP22", "5.30 PM"),    # evening time label
    ("A27", None),          # notes row
    ("A28", "Lecturers & Tutors"),  # key heading
]


@pytest.mark.parametrize("coord,value", STATIC_STYLE_CELLS)
def test_static_cell_style_and_value_preserved(db, template_ws, coord, value):
    ws = export(db)
    assert style_sig(ws[coord]) == style_sig(template_ws[coord]), f"{coord} style"
    if value is not None:
        assert ws[coord].value == value


def test_title_style_preserved_value_changed(db, template_ws):
    ws = export(db, "My Title")
    # Title style is the template's; only the value changes.
    assert style_sig(ws["B2"]) == style_sig(template_ws["B2"])
    assert ws["B2"].value == "My Title"


def test_version_style_preserved_value_changed(db, template_ws):
    ws = export(db)
    assert style_sig(ws["K29"]) == style_sig(template_ws["K29"])
    assert ws["K29"].value == "Version 1"


def test_key_body_style_matches_template(db, template_ws):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db, first="Steve", last="Chavura")
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    ws = export(db)
    # A generated key entry inherits the template key-cell style (A29).
    assert ws["A29"].value.startswith("SC:")
    assert style_sig(ws["A29"]) == style_sig(template_ws["A29"])


# ---------------------------------------------------------------------------
# Dynamic class style parity (copied from template subject exemplars)
# ---------------------------------------------------------------------------


# subject prefix -> (unit code, template exemplar cell)
SUBJECT_EXEMPLARS = {
    "HIS": ("HIS101", "C6"),
    "THE": ("THE202", "B6"),
    "LIT": ("LIT303", "O8"),
    "PHI": ("PHI204", "G16"),
    "GRE": ("GRE202", "AK6"),
    "LAN": ("LAN102", "M6"),
}


@pytest.mark.parametrize("prefix", list(SUBJECT_EXEMPLARS))
def test_class_cell_copies_template_subject_style(db, template_ws, prefix):
    code, exemplar = SUBJECT_EXEMPLARS[prefix]
    rooms = base_rooms(db)
    unit = make_unit(db, code=code)
    lec = make_lecturer(db)
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    ws = export(db)
    # B6 is the painted class cell; its full style equals the template exemplar.
    assert style_sig(ws["B6"]) == style_sig(template_ws[exemplar]), prefix


def test_class_cells_use_template_theme_fills_not_app_tokens(db):
    """History cells must use the template's theme-based subject fill (theme9
    tint 0.6) with default (automatic) text — not the old frontend subject token
    (F7E5D4 fill / coloured text)."""
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db)
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    ws = export(db)
    cell = ws["B6"]
    fg = cell.fill.fgColor
    assert fg.theme == 9 and round(fg.tint, 4) == 0.6
    assert cell.font.name == "Trebuchet MS" and cell.font.size == 10 and cell.font.b
    # Default/automatic text colour, not a subject-coloured RGB.
    assert cell.font.color is None
    assert cell.alignment.vertical == "top" and bool(cell.alignment.wrap_text)


def test_unknown_subject_uses_default_class_style(db, template_ws):
    rooms = base_rooms(db)
    unit = make_unit(db, code="SCI101")  # no template Science exemplar
    lec = make_lecturer(db)
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    ws = export(db)
    cell = ws["B6"]
    # Neutral white default class cell (template K6 exemplar's fill + font),
    # with the uniform medium box border.
    assert cell.fill.fgColor.theme == template_ws["K6"].fill.fgColor.theme == 0
    assert round(cell.fill.fgColor.tint or 0.0, 4) == 0.0
    assert cell.font.name == "Trebuchet MS" and cell.font.size == 10
    for side in (cell.border.top, cell.border.left, cell.border.right):
        assert side.style == "medium"


def test_multi_slot_class_styles_every_row(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db)
    s = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=2)
    make_assignment(db, "a1", s, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])

    ws = export(db)
    # s1+s2 => one merged block over rows 6..9 (both two-row bands), so the
    # subject fill on the anchor visually covers every row. The intermediate
    # band must NOT remain a separate merge or grey gap.
    ranges = merges(ws)
    assert "B6:B9" in ranges
    assert "B8:B9" not in ranges  # second band is absorbed into the single block
    fg = ws["B6"].fill.fgColor
    assert fg.theme == 9 and round(fg.tint, 4) == 0.6
    assert ws["B6"].value == "HIS101 Lecture (SC)"


# ---------------------------------------------------------------------------
# Block style parity (template blocked/static-event styling)
# ---------------------------------------------------------------------------


def test_named_block_copies_template_event_style(db, template_ws):
    rooms = base_rooms(db)
    make_block(
        db, "g1", "Chapel", BlockColour.GOLD,
        [(AvailabilityDay.MONDAY, AvailabilitySlot.S5, rooms["PDS"])],
    )
    ws = export(db)
    cell = ws["B16"]
    # Gold event fill + maroon text, copied from the template gold event (Z22).
    assert cell.fill.fgColor.rgb == template_ws["Z22"].fill.fgColor.rgb == "FFCC9900"
    assert cell.font.color.rgb == template_ws["Z22"].font.color.rgb == "FF660033"
    # Block cells carry the event alignment (center/center), not the class one.
    assert cell.alignment.horizontal == "center" and cell.alignment.vertical == "center"


def test_unnamed_block_uses_template_grey_blocked_style(db):
    rooms = base_rooms(db)
    make_block(
        db, "g1", None, None,
        [(AvailabilityDay.MONDAY, AvailabilitySlot.S5, rooms["PDS"])],
    )
    ws = export(db)
    assert ws["B16"].value is None
    # Grey "blocked" fill = the template's empty-cell theme fill (theme0 -0.15).
    fg = ws["B16"].fill.fgColor
    assert fg.theme == 0 and round(fg.tint, 4) == -0.15


# ---------------------------------------------------------------------------
# Placement stability — styling fixes introduce no placement change
# ---------------------------------------------------------------------------


def test_placements_unchanged_by_styling(db):
    rooms = base_rooms(db)
    unit = make_unit(db, code="HIS101")
    lec = make_lecturer(db, first="Steve", last="Chavura")
    s1 = make_session(db, "s1", unit, lec, SessionType.LECTURE, duration=1)
    s2 = make_session(db, "s2", unit, lec, SessionType.LECTURE, duration=1)
    make_assignment(db, "a1", s1, AvailabilityDay.MONDAY, AvailabilitySlot.S1, rooms["PDS"])
    make_assignment(db, "a2", s2, AvailabilityDay.FRIDAY, AvailabilitySlot.S7, rooms["JTW"])

    ws = export(db)
    assert ws["B6"].value == "HIS101 Lecture (SC)"     # Monday PDS s1
    assert ws["AO20"].value == "HIS101 Lecture (SC)"   # Friday JTW s7
