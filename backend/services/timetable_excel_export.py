"""Timetable Excel export service (Units 93, 96).

Renders the current *saved* timetable state into the fixed Campion timetable
``.xlsx`` template. The static template (``export_templates/
campion_timetable_export_template.xlsx``) owns all layout/styling; this service
only mutates the title, version, classes, blocks and the lecturer/tutor key,
then returns an in-memory workbook stream. Nothing is written to the database
or to disk — exports are never persisted.

Styling parity (Unit 96): class and block cells are painted by applying the
**template-derived NamedStyles** baked into the template by
``export_templates/build_template.py`` (copied verbatim from the source sheet's
own class/event exemplar cells). This module never constructs subject fills,
fonts, or borders from approximated RGB; it only chooses which template style
to apply per subject/block colour and writes the dynamic value. Static template
content (headers, time labels, ``Mass/Lunch``, ``FORMAL HALL``, notes, key,
version style) is left untouched.

Boundary note: choosing a class style derives a subject from the unit code
prefix. The subject parser is otherwise frontend-only, so this module keeps a
small, export-scoped subject -> template-style map purely for rendering the
Excel artifact (see ``SUBJECT_PREFIXES``). It does not introduce a general
backend subject parser.
"""
from __future__ import annotations

import io
import re
from copy import copy
from datetime import date
from pathlib import Path

import structlog
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session as DBSession, selectinload

from api.errors import AppError
from models.assignment import TimetableAssignment
from models.room import Room
from models.session import Session, SessionType
from models.timetable_block import TimetableBlockCell, TimetableBlockGroup
from models.unit import Unit

logger = structlog.get_logger(__name__)

# Repo-owned static template, derived once from the source workbook by
# export_templates/build_template.py.
EXPORT_TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent
    / "export_templates"
    / "campion_timetable_export_template.xlsx"
)

# --- Fixed sheet mapping (must mirror export_templates/build_template.py) -----

# Fixed room order per day, left to right within each day block.
ROOM_ORDER: list[str] = ["PDS", "L1.05", "Bromley", "L1.08", "Dawson", "L1.10", "L1.12", "JTW"]
ROOMS_PER_DAY = len(ROOM_ORDER)

# Days map left to right; each day is an 8-column block (Monday starts at B=2).
DAY_ORDER: list[str] = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
_DAY_START_COL: dict[str, int] = {
    day: 2 + index * ROOMS_PER_DAY for index, day in enumerate(DAY_ORDER)
}

# App slots map to the main visible slot rows; each slot fills a two-row band.
ORDERED_SLOTS: list[str] = ["s1", "s2", "s3", "s4", "s5", "s6", "s7"]
_SLOT_TOP_ROW: dict[str, int] = {
    "s1": 6, "s2": 8, "s3": 10, "s4": 14, "s5": 16, "s6": 18, "s7": 20,
}

TITLE_CELL = "B2"
VERSION_CELL = "K29"
STATIC_VERSION = "Version 1"

# Lecturer/tutor key area: two columns (A, D) starting at row 29. The template
# styles five key rows per column (29-33); overflow rows copy that style.
KEY_COLUMNS = ["A", "D"]
KEY_START_ROW = 29
KEY_TEMPLATE_LAST_ROW = 33
KEY_CLEAR_ROWS = range(29, 41)

# --- Styling (template-derived NamedStyles, baked by build_template.py) ------

# Subject class styles keyed by unit-code prefix. Each name resolves to a
# NamedStyle copied verbatim from the template's own class exemplar cell.
# Unknown/legacy prefixes (e.g. SCI) fall back to the neutral default class
# style. Must mirror export_templates/build_template.py.
SUBJECT_STYLE_PREFIX = "tt_class_"
SUBJECT_PREFIXES: frozenset[str] = frozenset({"HIS", "THE", "LIT", "PHI", "GRE", "LAN"})
DEFAULT_CLASS_STYLE = "tt_class_default"

# Block styles by app block colour; resolve to the template's blocked/static-
# event styling (gold/blue/maroon events; grey "blocked" for unnamed blocks).
BLOCK_STYLE: dict[str | None, str] = {
    "gold": "tt_block_gold",
    "light_blue": "tt_block_blue",
    "light_pink": "tt_block_pink",
    None: "tt_block_unnamed",
}

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _class_style_name(unit_code: str) -> str:
    prefix = unit_code[:3].upper()
    if prefix in SUBJECT_PREFIXES:
        return SUBJECT_STYLE_PREFIX + prefix
    return DEFAULT_CLASS_STYLE


# --- Mapping helpers ---------------------------------------------------------


def _room_index(room_name: str) -> int:
    try:
        return ROOM_ORDER.index(room_name)
    except ValueError:
        raise AppError(
            "export_room_not_in_template",
            f"Room '{room_name}' is not part of the fixed timetable template.",
            status_code=422,
        )


def cell_anchor(day: str, room_name: str, slot: str) -> tuple[int, int]:
    """Return the (row, col) of the top-left cell for a day/room/slot."""
    if day not in _DAY_START_COL:
        raise AppError(
            "export_invalid_assignment_shape",
            f"Unknown timetable day '{day}'.",
            status_code=422,
        )
    if slot not in _SLOT_TOP_ROW:
        raise AppError(
            "export_invalid_assignment_shape",
            f"Unknown timetable slot '{slot}'.",
            status_code=422,
        )
    col = _DAY_START_COL[day] + _room_index(room_name)
    return _SLOT_TOP_ROW[slot], col


def _slot_row_span(start_slot: str, duration: int) -> tuple[int, int]:
    """Return (top_row, bottom_row) covering all two-row bands of a session."""
    start_index = ORDERED_SLOTS.index(start_slot)
    end_index = start_index + duration - 1
    if end_index >= len(ORDERED_SLOTS):
        raise AppError(
            "export_invalid_assignment_shape",
            f"Session starting at {start_slot} with duration {duration} runs off the timetable.",
            status_code=422,
        )
    last_slot = ORDERED_SLOTS[end_index]
    return _SLOT_TOP_ROW[start_slot], _SLOT_TOP_ROW[last_slot] + 1


# --- Workbook painting -------------------------------------------------------


def _unmerge_region(ws: Worksheet, top: int, left: int, bottom: int, right: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        if not (rng.max_row < top or rng.min_row > bottom
                or rng.max_col < left or rng.min_col > right):
            ws.unmerge_cells(str(rng))


def _paint(
    ws: Worksheet,
    top: int,
    left: int,
    bottom: int,
    right: int,
    style_name: str,
    value: str | None,
) -> None:
    """Apply a template-derived NamedStyle to a rectangle, merge it, and write
    the value at the top-left cell.

    The style (fill/font/border/alignment) is copied verbatim from the
    template's own class/event exemplar cells (baked into the template by
    build_template.py); this function never constructs approximated styles.
    """
    if top < 6 or bottom > 21 or 12 <= top <= 13 or 12 <= bottom <= 13:
        # App slot bands never touch the static Mass/Lunch rows (12-13); guard
        # against a malformed placement clobbering static template content.
        raise AppError(
            "export_overlaps_static_content",
            "A timetable placement unexpectedly overlaps static template content.",
            status_code=409,
        )

    _unmerge_region(ws, top, left, bottom, right)
    for row in range(top, bottom + 1):
        for col in range(left, right + 1):
            ws.cell(row=row, column=col).style = style_name
    if top != bottom or left != right:
        ws.merge_cells(start_row=top, start_column=left, end_row=bottom, end_column=right)
    ws.cell(row=top, column=left).value = value


# --- Labels ------------------------------------------------------------------


def lecturer_initials(first_name: str, last_name: str) -> str:
    """First-name initial + last-name initial, uppercased, ignoring title."""
    return f"{first_name.strip()[:1]}{last_name.strip()[:1]}".upper()


def _require_lecturer(session: Session) -> object:
    lecturer = session.lecturer
    if lecturer is None:
        raise AppError(
            "export_session_missing_lecturer",
            "A scheduled session is missing its lecturer and cannot be exported.",
            status_code=422,
        )
    return lecturer


def _session_label(session: Session, tutorial_letter: str | None) -> str:
    lecturer = _require_lecturer(session)
    initials = lecturer_initials(lecturer.first_name, lecturer.last_name)
    code = session.unit.code
    if session.session_type == SessionType.TUTORIAL:
        suffix = f" {tutorial_letter}" if tutorial_letter else ""
        return f"{code} Tutorial{suffix} ({initials})"
    return f"{code} Lecture ({initials})"


# --- Tutorial lettering ------------------------------------------------------


def _tutorial_letters(assignments: list[TimetableAssignment]) -> dict[str, str]:
    """Assign export-only tutorial letters per unit, in timetable order.

    Order: day (Mon-Fri), start slot (s1-s7), room (fixed template order),
    then stable session id as the final tie-breaker. Only exported tutorials
    consume letters.
    """
    by_unit: dict[str, list[TimetableAssignment]] = {}
    for a in assignments:
        if a.session.session_type == SessionType.TUTORIAL:
            by_unit.setdefault(a.session.unit_id, []).append(a)

    letters: dict[str, str] = {}
    for unit_id, items in by_unit.items():
        ordered = sorted(
            items,
            key=lambda a: (
                DAY_ORDER.index(a.day.value),
                ORDERED_SLOTS.index(a.start_slot.value),
                _room_index(a.room.name),
                a.session_id,
            ),
        )
        for index, a in enumerate(ordered):
            letters[a.session_id] = chr(ord("A") + index)
    return letters


# --- Block rectangle merging -------------------------------------------------


def _block_rectangles(
    cells: list[tuple[str, str, str]]
) -> list[tuple[int, int, int, int]]:
    """Group a block group's (day, slot, room) cells into merged rectangles.

    Mirrors the frontend rectangle merging: per (day, room) find row-contiguous
    slot runs, then merge runs of the same shape across adjacent room columns.
    Runs never merge across the static Mass/Lunch gap (top rows differ by 2).
    """
    # Per (day, room_index): the set of slot top-rows occupied.
    by_day_room: dict[tuple[str, int], list[int]] = {}
    for day, slot, room_name in cells:
        anchor_row, col = cell_anchor(day, room_name, slot)
        by_day_room.setdefault((day, _room_index(room_name)), []).append(anchor_row)

    # Build vertical runs keyed by (day, top_row, bottom_row) -> set of room indices.
    runs: dict[tuple[str, int, int], set[int]] = {}
    for (day, room_index), rows in by_day_room.items():
        for top, bottom in _contiguous_row_runs(sorted(set(rows))):
            runs.setdefault((day, top, bottom), set()).add(room_index)

    rectangles: list[tuple[int, int, int, int]] = []
    for (day, top, bottom), room_indices in runs.items():
        start_col = _DAY_START_COL[day]
        for first, last in _contiguous_int_runs(sorted(room_indices)):
            rectangles.append((top, start_col + first, bottom, start_col + last))
    return rectangles


def _contiguous_row_runs(top_rows: list[int]) -> list[tuple[int, int]]:
    """Collapse sorted slot top-rows into (top, bottom) runs, split at the
    lunch gap (adjacent bands differ by exactly 2 rows)."""
    runs: list[tuple[int, int]] = []
    run_start = top_rows[0]
    prev = top_rows[0]
    for row in top_rows[1:]:
        if row == prev + 2:
            prev = row
            continue
        runs.append((run_start, prev + 1))
        run_start = prev = row
    runs.append((run_start, prev + 1))
    return runs


def _contiguous_int_runs(values: list[int]) -> list[tuple[int, int]]:
    runs: list[tuple[int, int]] = []
    start = prev = values[0]
    for v in values[1:]:
        if v == prev + 1:
            prev = v
            continue
        runs.append((start, prev))
        start = prev = v
    runs.append((start, prev))
    return runs


# --- Lecturer/tutor key ------------------------------------------------------


def _write_lecturer_key(ws: Worksheet, assignments: list[TimetableAssignment]) -> None:
    # Unique lecturers used by exported saved sessions, keyed by lecturer id.
    by_id: dict[str, tuple[str, str]] = {}
    for a in assignments:
        lecturer = _require_lecturer(a.session)
        initials = lecturer_initials(lecturer.first_name, lecturer.last_name)
        display = f"{lecturer.title.value} {lecturer.first_name} {lecturer.last_name}"
        by_id[lecturer.id] = (initials, display)

    entries = sorted(by_id.values(), key=lambda e: (e[0], e[1]))

    # Clear the key value cells (template key style preserved), then write
    # balanced columns. Written cells inherit the template key-area style; any
    # overflow rows below the styled key block copy the style of their column's
    # first key cell so the look stays template-faithful.
    for col in KEY_COLUMNS:
        for row in KEY_CLEAR_ROWS:
            ws[f"{col}{row}"] = None

    per_column = (len(entries) + 1) // 2  # column A holds the first half
    for index, (initials, display) in enumerate(entries):
        col = KEY_COLUMNS[0] if index < per_column else KEY_COLUMNS[1]
        offset = index if index < per_column else index - per_column
        row = KEY_START_ROW + offset
        cell = ws[f"{col}{row}"]
        cell.value = f"{initials}: {display}"
        if row > KEY_TEMPLATE_LAST_ROW:
            cell._style = copy(ws[f"{col}{KEY_START_ROW}"]._style)


# --- Public API --------------------------------------------------------------


def _load_template() -> Worksheet:
    if not EXPORT_TEMPLATE_PATH.exists():
        # Never leak the filesystem path to the caller.
        logger.error("export_template_missing", configured=str(EXPORT_TEMPLATE_PATH))
        raise AppError(
            "export_template_missing",
            "The timetable export template is unavailable.",
            status_code=500,
        )
    wb = load_workbook(EXPORT_TEMPLATE_PATH)
    return wb


def generate_timetable_export(db: DBSession, title: str) -> io.BytesIO:
    """Render the saved timetable into the static template and return a stream.

    Reads only canonical saved data (assignments + blocks); does not mutate any
    assignment state and does not persist the generated workbook.
    """
    cleaned = title.strip()
    if not cleaned:
        raise AppError(
            "invalid_export_title",
            "A non-empty timetable title is required.",
            status_code=422,
        )

    wb = _load_template()
    ws = wb.worksheets[0]

    ws[TITLE_CELL] = cleaned
    ws[VERSION_CELL] = STATIC_VERSION

    assignments = (
        db.query(TimetableAssignment)
        .options(
            selectinload(TimetableAssignment.session).selectinload(Session.lecturer),
            selectinload(TimetableAssignment.session).selectinload(Session.unit),
            selectinload(TimetableAssignment.room),
        )
        .all()
    )

    tutorial_letters = _tutorial_letters(assignments)

    # Write classes (saved scheduled assignments only).
    for a in assignments:
        session = a.session
        room_name = a.room.name
        top, col = cell_anchor(a.day.value, room_name, a.start_slot.value)
        _, bottom = _slot_row_span(a.start_slot.value, session.duration)
        style_name = _class_style_name(session.unit.code)
        label = _session_label(session, tutorial_letters.get(a.session_id))
        _paint(ws, top, col, bottom, col, style_name, label)

    _write_blocks(db, ws)
    _write_lecturer_key(ws, assignments)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    logger.info(
        "timetable_export_generated",
        assignment_count=len(assignments),
    )
    return stream


def _write_blocks(db: DBSession, ws: Worksheet) -> None:
    rooms_by_id: dict[str, str] = {r.id: r.name for r in db.query(Room).all()}
    groups = (
        db.query(TimetableBlockGroup)
        .options(selectinload(TimetableBlockGroup.cells))
        .all()
    )
    for group in groups:
        cells: list[tuple[str, str, str]] = []
        for cell in group.cells:
            room_name = rooms_by_id.get(cell.room_id)
            if room_name is None or room_name not in ROOM_ORDER:
                raise AppError(
                    "export_room_not_in_template",
                    "A timetable block uses a room that is not part of the fixed template.",
                    status_code=422,
                )
            cells.append((cell.day.value, cell.slot.value, room_name))
        if not cells:
            continue
        colour = group.colour.value if group.colour is not None else None
        style_name = BLOCK_STYLE.get(colour, BLOCK_STYLE[None])
        value = group.name if group.name else None
        for top, left, bottom, right in _block_rectangles(cells):
            _paint(ws, top, left, bottom, right, style_name, value)


def export_filename(title: str, today: date) -> str:
    """Safe slug of the title plus the current date, e.g.
    ``semester-2-2026-timetable-2026-06-20.xlsx``."""
    slug = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-") or "timetable"
    return f"{slug}-{today.isoformat()}.xlsx"
