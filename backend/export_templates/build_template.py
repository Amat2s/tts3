"""Build the static Campion timetable export template (Units 93, 96, 97).

This is a one-off, repo-owned build step. It derives the committed static
template ``campion_timetable_export_template.xlsx`` from the **real pristine**
Campion source workbook (``assets/excel/Timetable S2 2025 DRAFT.xlsx``), so the
export service can load a clean, stable layout and only mutate the title,
version, lecturer key, classes and blocks at request time.

Unit 97 (template fidelity) rebuilt this from the correct source sheet after
Units 93/96 were found to have been built from a *stripped* ``wijmo.xlsx``
re-export that had lost the real template's page setup, tab colour, sheet zoom,
fine-grained row heights, default column model, per-day background shading, and
bold theme-coloured borders. Concretely this build:

* selects the canonical sheet **by name** ``S2, 2025 Timetable `` (note the
  trailing space) and removes the other two sheets, so the output is a
  single-sheet workbook;
* keeps the blank template a **verbatim blanked copy** of that sheet: column
  widths (including the columns the template leaves default/unset), row heights,
  sheet-format properties, page setup, margins, tab colour, sheet view/zoom, and
  every static merged range, static value, and static style are preserved
  untouched (day/room headers, time labels, the static ``Mass/Lunch`` row, the
  evening rows incl. the static ``FORMAL HALL`` banner, the notes area, and the
  lecturer/tutor key layout);
* clears only the dynamic class/event **values** from the s1-s7 grid bands and
  resets each formerly populated grid cell to the empty look of **its own day
  block** — copied verbatim from a real empty 2-row band in the *same column*
  (so Tuesday/Thursday stay white ``theme0`` tint 0.0, Monday/Wednesday/Friday
  stay grey ``theme0`` tint -0.15, and the template's medium row borders and
  dark day-boundary/interior verticals are preserved, never reconstructed from
  approximated RGB), re-merging each room x slot into its 2-row band;
* bakes **template-derived NamedStyles** for each subject class and each block
  colour, copied directly from the template's own class/event exemplar cells
  (theme-based subject fills; gold/blue/maroon event fills + text), so the
  export applies real template styles instead of constructing new ones — block
  cells carry the template's *event* alignment (not the class alignment);
* clears the old lecturer/tutor key names (the export regenerates them) while
  keeping the key-area styling, and sets the version text to ``Version 1``.

Run from the ``backend`` directory:

    python export_templates/build_template.py

The generated ``.xlsx`` is committed to the repository; this script only needs
to be re-run if the source layout changes.
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, NamedStyle

# --- Layout constants (must mirror services/timetable_excel_export.py) -------

# Day blocks are 8 room columns wide, left to right Monday..Friday.
DAY_START_COLS = [2, 10, 18, 26, 34]  # B, J, R, Z, AH
ROOMS_PER_DAY = 8

# Each app slot occupies a two-row visual band; value lives in the top row.
SLOT_TOP_ROWS = [6, 8, 10, 14, 16, 18, 20]  # s1..s7
LAST_COL = DAY_START_COLS[-1] + ROOMS_PER_DAY - 1  # AO = 41
GRID_COLS = [c for start in DAY_START_COLS for c in range(start, start + ROOMS_PER_DAY)]

# Lecturer/tutor key value cells to clear (layout + style preserved).
KEY_VALUE_CELLS = [
    f"{col}{row}" for col in ("A", "D") for row in range(29, 34)
]

VERSION_CELL = "K29"
STATIC_VERSION = "Version 1"

# Canonical source: the real pristine template, S2 sheet (trailing space).
SOURCE = Path("assets/excel/Timetable S2 2025 DRAFT.xlsx")
SOURCE_SHEET = "S2, 2025 Timetable "
OUTPUT = Path("export_templates/campion_timetable_export_template.xlsx")

# --- Template style exemplars (cells copied verbatim from the source sheet) ---
#
# These are the canonical example cells the template itself uses; their styles
# are copied rather than re-approximated. Subject codes are keyed by unit-code
# prefix and mirror services/timetable_excel_export.py. The new sheet uses
# theme-based subject fills (e.g. Theology theme5/0.6, History theme9/0.6,
# Literature theme2/-0.25).

EMPTY_EXEMPLAR = "D6"  # a real grey "empty" grid cell (Monday, theme0 tint -0.15)

SUBJECT_EXEMPLARS = {
    "HIS": "C6",   # History     (theme9 tint 0.6)
    "THE": "B6",   # Theology    (theme5 tint 0.6)
    "LIT": "O8",   # Literature  (theme2 tint -0.25)
    "PHI": "G16",  # Philosophy  (theme4 tint 0.6)
    "GRE": "AK6",  # Greek       (theme6 tint 0.6)
    "LAN": "M6",   # Latin/Languages (theme9 tint 0.8)
}
# Unknown/legacy prefixes (e.g. SCI) fall back to a neutral white class cell.
DEFAULT_CLASS_EXEMPLAR = "K6"  # a real white grid cell (theme0 tint 0.0)

# Static-event exemplars used to derive block styling (the app's blocks reuse
# the template's blocked/static-event look rather than frontend block tokens).
GOLD_EVENT_EXEMPLAR = "Z22"    # FORMAL HALL: gold cc9900 fill, maroon 660033 text
BLUE_EVENT_EXEMPLAR = "R10"    # Augustine Academy: blue 0070c0 fill, black text
MAROON_EVENT_EXEMPLAR = "B12"  # Mass/Lunch: maroon 660033 fill, gold cc9900 text

# Named styles the export applies. Must match the names in
# services/timetable_excel_export.py.
SUBJECT_STYLE_PREFIX = "tt_class_"
DEFAULT_CLASS_STYLE = "tt_class_default"
BLOCK_STYLE = {
    "gold": "tt_block_gold",
    "light_blue": "tt_block_blue",
    "light_pink": "tt_block_pink",
    None: "tt_block_unnamed",
}


def _class_named_style(name: str, exemplar) -> NamedStyle:
    """A NamedStyle that copies a template class cell's style verbatim."""
    ns = NamedStyle(name=name)
    ns.font = copy(exemplar.font)
    ns.fill = copy(exemplar.fill)
    ns.border = copy(exemplar.border)
    ns.alignment = copy(exemplar.alignment)
    ns.number_format = exemplar.number_format
    return ns


def _block_named_style(name: str, event_exemplar, class_exemplar) -> NamedStyle:
    """A block NamedStyle reusing a template event's fill + text colour and the
    template event's alignment, with the class cell's readable 10pt bold font
    and clean medium-box border (copied verbatim, never reconstructed)."""
    ns = NamedStyle(name=name)
    base = class_exemplar.font
    ns.font = Font(
        name=base.name, size=base.size, bold=True,
        color=copy(event_exemplar.font.color),
    )
    ns.fill = copy(event_exemplar.fill)
    ns.border = copy(class_exemplar.border)
    # Block cells carry the template's static-event alignment (defect 4), which
    # differs from the class alignment (e.g. FORMAL HALL is center/center).
    ns.alignment = copy(event_exemplar.alignment)
    ns.number_format = class_exemplar.number_format
    return ns


def _register_named_styles(ws, wb) -> None:
    class_exemplar = ws[SUBJECT_EXEMPLARS["HIS"]]
    for prefix, coord in SUBJECT_EXEMPLARS.items():
        wb.add_named_style(_class_named_style(SUBJECT_STYLE_PREFIX + prefix, ws[coord]))
    wb.add_named_style(_class_named_style(DEFAULT_CLASS_STYLE, ws[DEFAULT_CLASS_EXEMPLAR]))

    # Blocks: gold / blue / maroon(pink) event styling, plus a grey "blocked"
    # style for unnamed blocks (the template's own empty-cell look, verbatim).
    wb.add_named_style(_block_named_style(
        BLOCK_STYLE["gold"], ws[GOLD_EVENT_EXEMPLAR], class_exemplar,
    ))
    wb.add_named_style(_block_named_style(
        BLOCK_STYLE["light_blue"], ws[BLUE_EVENT_EXEMPLAR], class_exemplar,
    ))
    wb.add_named_style(_block_named_style(
        BLOCK_STYLE["light_pink"], ws[MAROON_EVENT_EXEMPLAR], class_exemplar,
    ))
    wb.add_named_style(_class_named_style(BLOCK_STYLE[None], ws[EMPTY_EXEMPLAR]))


def _is_empty_grid_cell(cell) -> bool:
    """True if a grid cell is a real empty band anchor: no value and a theme0
    (grey/white) empty fill, as opposed to a themed class fill or an rgb event
    fill. Merge continuations (patternType None) are not band anchors."""
    if cell.value is not None:
        return False
    fill = cell.fill
    return fill.patternType == "solid" and isinstance(fill.fgColor.theme, int) and fill.fgColor.theme == 0


def _empty_band_style(ws) -> dict[int, tuple[tuple, tuple]]:
    """For each grid column, capture the (top-cell, bottom-cell) style of a real
    empty 2-row band in that column, so blanked cells can be reset to their own
    day block's empty look verbatim. Captured before any mutation.

    Style tuples are (fill, font, alignment, border) copies. If a column is
    fully populated (no empty band of its own) it falls back to another empty
    column in the same day block.
    """
    def capture(col: int):
        for top in SLOT_TOP_ROWS:
            anchor = ws.cell(row=top, column=col)
            if _is_empty_grid_cell(anchor):
                bottom = ws.cell(row=top + 1, column=col)
                return (
                    (copy(anchor.fill), copy(anchor.font), copy(anchor.alignment), copy(anchor.border)),
                    (copy(bottom.fill), copy(bottom.font), copy(bottom.alignment), copy(bottom.border)),
                )
        return None

    exemplars: dict[int, tuple[tuple, tuple]] = {}
    for col in GRID_COLS:
        style = capture(col)
        if style is None:
            # Fall back to any empty column in the same day block.
            start = next(s for s in DAY_START_COLS if s <= col < s + ROOMS_PER_DAY)
            for sibling in range(start, start + ROOMS_PER_DAY):
                style = capture(sibling)
                if style is not None:
                    break
        if style is None:  # pragma: no cover - the real template always has empties
            raise RuntimeError(f"No empty grid band found for column {col} or its day block.")
        exemplars[col] = style
    return exemplars


def _grid_merges(ws):
    grid_rows = {r for top in SLOT_TOP_ROWS for r in (top, top + 1)}
    for rng in list(ws.merged_cells.ranges):
        rows = set(range(rng.min_row, rng.max_row + 1))
        if rows <= grid_rows and rng.min_col >= 2 and rng.max_col <= LAST_COL:
            yield rng


def _populated_bands(ws) -> set[tuple[int, int]]:
    """Set of (column, band_top) grid bands that hold a dynamic class/event
    value (i.e. must be blanked). A band is populated when the merge covering it
    has an anchor that is not a real empty cell (has a value or a non-empty
    themed/rgb fill). Already-empty bands are excluded so their own verbatim
    fill/border are preserved."""
    populated: set[tuple[int, int]] = set()
    for rng in _grid_merges(ws):
        anchor = ws.cell(row=rng.min_row, column=rng.min_col)
        if _is_empty_grid_cell(anchor):
            continue
        for col in range(rng.min_col, rng.max_col + 1):
            for top in SLOT_TOP_ROWS:
                if rng.min_row <= top and top + 1 <= rng.max_row:
                    populated.add((col, top))
    return populated


def _blank_grid(ws) -> None:
    """Clear the dynamic s1-s7 grid. Formerly-populated bands are reset to their
    own day block's empty look (copied verbatim from a real empty band in the
    same column, so Tuesday/Thursday stay white and Mon/Wed/Fri grey, with the
    template's medium row borders and day-boundary verticals). Already-empty
    bands keep their own verbatim fill/border. Every room x slot is re-merged
    into its 2-row band."""
    exemplars = _empty_band_style(ws)
    populated = _populated_bands(ws)

    for rng in _grid_merges(ws):
        ws.unmerge_cells(str(rng))

    for col in GRID_COLS:
        (tf, tfont, talign, tborder), (bf, bfont, balign, bborder) = exemplars[col]
        for top in SLOT_TOP_ROWS:
            if (col, top) in populated:
                anchor = ws.cell(row=top, column=col)
                anchor.value = None
                anchor.fill = copy(tf)
                anchor.font = copy(tfont)
                anchor.alignment = copy(talign)
                anchor.border = copy(tborder)

                bottom = ws.cell(row=top + 1, column=col)
                bottom.value = None
                bottom.fill = copy(bf)
                bottom.font = copy(bfont)
                bottom.alignment = copy(balign)
                bottom.border = copy(bborder)

            ws.merge_cells(start_row=top, start_column=col, end_row=top + 1, end_column=col)


def build() -> Path:
    wb = openpyxl.load_workbook(SOURCE)

    # One-sheet workbook only: keep the canonical S2 layout sheet (by name, NOT
    # worksheets[0], which is the generic "Timetable Template" sheet), drop rest.
    keep = wb[SOURCE_SHEET]
    for ws in list(wb.worksheets):
        if ws is not keep:
            wb.remove(ws)
    ws = keep

    # Bake template-derived class/block styles BEFORE blanking the grid, while
    # the exemplar class/event cells still carry their original styles.
    _register_named_styles(ws, wb)

    _blank_grid(ws)

    # Static version text (style preserved).
    ws[VERSION_CELL] = STATIC_VERSION

    # Clear old lecturer/tutor key names; the export regenerates them. Keep the
    # cell styling so generated entries inherit the key-area look.
    for coord in KEY_VALUE_CELLS:
        ws[coord].value = None

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
